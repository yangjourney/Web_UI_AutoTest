#encoding=utf-8
from openpyxl import *
from openpyxl.styles import Font
import logging
from Util.KeyWordDriven.KeyWorldTool.FormatTime import *

class ParseExcel(object):
    def __init__(self,excel_file_path):
        self.excel_file_path=excel_file_path
        self.workbook=load_workbook(excel_file_path)
        self.font=Font(color=None)
        self.colorDict={"red":'FFFF3030',"green":'FF008B00'}
        #self.sheet=self.get_sheet_by_index(0)
        self.sheet=None

    #设置当前要操作的sheet对象，使用index来获取相应的sheet
    def set_sheet_by_index(self,sheet_index):
        self.log_info('set_sheet_by_index  sheet_index='+sheet_index)
        self.sheet = self.get_sheet_by_index(sheet_index)

    # 设置当前要操作的sheet对象，使用sheet名称来获取相应的sheet
    def set_sheet_by_name(self,sheet_name):
        self.log_info('set_sheet_by_name sheet_name='+sheet_name)
        self.sheet = self.workbook.get_sheet_by_name(sheet_name)

    #获取当前默认sheet的名字
    def get_default_name(self):
        self.log_info('get_default_name ')
        return self.sheet.title

    #通过sheet名称获取sheet对象
    def get_sheet_by_name(self,sheet_name):
        self.log_info('get_sheet_by_name sheet_name='+sheet_name)
        self.sheet = self.workbook[sheet_name]
        return self.workbook[sheet_name]

    # 通过sheet 索引获取sheet对象
    def get_sheet_by_index(self,sheet_index):
        self.log_info('get_sheet_by_index  sheet_index='+sheet_index)
        sheet_name=self.workbook.get_sheet_names()[sheet_index]
        self.sheet=self.get_sheet_by_name(sheet_name)
        return self.get_sheet_by_name(sheet_name)
    #获取全部sheet页名称
    def get_all_sheet_names(self):
        self.log_info('get_all_sheet_names')
        return self.workbook.sheetnames

    #获取默认sheet中最大的行数
    def get_max_row_no(self):
        self.log_info('get_max_row_no')
        return self.sheet.max_row

    #获取默认 sheet 的最大列数
    def get_max_col_no(self):
        self.log_info('get_max_col_no')
        return self.sheet.max_column

    #获取默认sheet的最小（起始）行号
    def get_min_row_no(self):
        self.log_info('get_min_row_no')
        return self.sheet.min_row

    # 获取默认sheet的最小（起始）列号
    def get_min_col_no(self):
        self.log_info('get_min_col_no')
        return self.sheet.min_column


    # 获取默认 sheet 的所有行对象
    def get_all_rows(self):
        self.log_info('get_all_rows')
        return self.sheet.rows

    #获取默认sheet中的所有列对象
    def get_all_cols(self):
        self.log_info('get_all_cols')
        return self.sheet.columns

    #从默认sheet中获取某一行，第一行从0开始
    def get_single_row(self,row_no):
        self.log_info('get_single_row  row_no='+row_no)
        return self.get_all_rows()[row_no]

    # #从默认sheet中获取某一列，第一列从0开始
    def get_single_col(self,col_no):
        self.log_info('get_single_col  col_no=' + col_no)
        return self.get_all_cols()[col_no]

    #从默认sheet中，通过行号和列号获取指定的单元格，注意行号和列号从1开始
    def get_cell(self,row_no,col_no):
        self.log_info('get_cell  row_no=' + row_no+', col_no='+col_no)
        return self.sheet.cell(row=row_no,column=col_no)

    # 从默认sheet中，通过行号和列号获取指定的单元格中的内容，注意行号和列号从1开始
    def get_cell_content(self,row_no,col_no):
        self.log_info('get_cell_content  row_no=' + row_no + ', col_no=' + col_no)
        return self.sheet.cell(row=row_no, column=col_no).value

    # 从默认sheet中，通过行号和列号向指定单元格中写入指定内容，注意行号和列号从1开始
    # 调用此方法的时候，excel不要处于打开状态
    def write_cell_content(self,row_no,col_no,content,font=None):
        self.log_info('write_cell_content  row_no=' + str(row_no) + ', col_no=' + str(col_no)+', content='+content)
        self.sheet.cell(row=row_no, column=col_no).value=content
        self.workbook.save(self.excel_file_path)

    # 从默认sheet中，通过行号和列号向指定单元格中写入当前日期，注意行号和列号从1开始
    #调用此方法的时候，excel不要处于打开状态
    def write_cell_current_time(self,row_no,col_no):
        date_time_chinese=time.strftime("%Y-%m-%d %H:%M:%S")
        self.log_info('write_cell_current_time  row_no=' + str(row_no) + ', col_no=' + str(col_no))
        self.sheet.cell(row=row_no, column=col_no).value =str(date_time_chinese)
        self.workbook.save(self.excel_file_path)
    #保存
    def workbook_save(self):
        self.log_info('workbook_save  ')
        self.workbook.save(self.excel_file_path)

    def log_info(self, info):
        # logging.info(str(info))
        # print(str(info))
        pass

    def log_error(self, info):
        logging.error(str(info))
        print(str(info))
#关键字驱动
#测试用例页面
case_no=0
case_name=1
case_describe=2
case_detail=3
case_NeedToDo=4
case_usetime=5
case_result=6

#测试步骤页面
step_no=0
step_describe=1
step_keyword=2
step_method=3
step_expression=4
step_value=5
step_time=6
step_result=7
step_err=8



